import io
import uuid
import openpyxl
from decimal import Decimal
from datetime import datetime, date
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from models.models import (
    LookupValue,
    Expense,
    Income,
    Liability,
    LiabilityPayment,
    Investment,
)
from seed_data import seed_user_data

class ImportService:
    @staticmethod
    def to_date(val):
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        return None

    @staticmethod
    def extract_day_from_string(due_date_str: str) -> int:
        if not due_date_str:
            return None
        import re
        match = re.search(r"\d+", str(due_date_str))
        if match:
            return int(match.group())
        return None

    @classmethod
    async def wipe_user_data(cls, db: AsyncSession, user_id: str):
        """Wipe all financial data for a user."""
        # Order matters for foreign keys if any, although here they mostly point to lookup_values
        await db.execute(delete(LiabilityPayment).where(LiabilityPayment.liability_id.in_(
            select(Liability.id).where(Liability.user_id == user_id)
        )))
        await db.execute(delete(Liability).where(Liability.user_id == user_id))
        await db.execute(delete(Expense).where(Expense.user_id == user_id))
        await db.execute(delete(Income).where(Income.user_id == user_id))
        await db.execute(delete(Investment).where(Investment.user_id == user_id))
        # Note: We don't wipe LookupValues as they are usually categories the user might want to keep,
        # but the migration script expects to ensure them. We'll keep them for now.
        await db.commit()

    @classmethod
    async def import_from_excel(cls, db: AsyncSession, user_id: str, file_content: bytes):
        """Import data from an Excel file content."""
        # 1. Wipe existing data
        await cls.wipe_user_data(db, user_id)

        # 2. Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

        # 3. Ensure lookup values exist
        await seed_user_data(db, user_id)
        
        # Get all lookups for this user
        result = await db.execute(select(LookupValue).where(LookupValue.user_id == user_id))
        all_lookups = result.scalars().all()
        lookup_dict = {}
        for lv in all_lookups:
            if lv.type not in lookup_dict:
                lookup_dict[lv.type] = {}
            lookup_dict[lv.type][lv.value] = lv.id

        def get_lookup_id(ltype, lval, default_val=None):
            if ltype not in lookup_dict:
                return None
            
            if lval in lookup_dict[ltype]:
                return lookup_dict[ltype][lval]
            elif default_val and default_val in lookup_dict[ltype]:
                return lookup_dict[ltype][default_val]

            # Try case-insensitive strip match
            for k, v in lookup_dict[ltype].items():
                if str(k).strip().lower() == str(lval).strip().lower():
                    return v
            return None

        account_icici_id = get_lookup_id("account", "ICICI Bank") or get_lookup_id("account", "Savings Account")

        # 4. Liabilities
        liabilities_map = {}
        ws = wb["Debts List"] if "Debts List" in wb.sheetnames else None
        if ws:
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                if not row or not any(row): continue
                loan_type = row[1]
                lender = row[2]
                if not loan_type and not lender: continue

                original_amount = row[3] if row[3] is not None else Decimal("0")
                interest_rate = row[4] if row[4] is not None else 0.0
                tenure_months = row[5] if row[5] is not None else 0
                emi = row[6] if row[6] is not None else Decimal("0")
                start_date = cls.to_date(row[0]) or date.today()
                due_day = cls.extract_day_from_string(row[11]) if len(row) > 11 else None

                type_id = get_lookup_id("liabilities_type", loan_type)
                lender_id = get_lookup_id("lender", lender)

                if type_id and lender_id:
                    lib = Liability(
                        id=uuid.uuid4(),
                        user_id=user_id,
                        liabilities_type_id=type_id,
                        lender_id=lender_id,
                        liability_name=f"{loan_type} - {lender}",
                        original_amount=Decimal(str(original_amount)),
                        interest_rate=float(interest_rate),
                        emi_amount=Decimal(str(emi)),
                        start_date=start_date,
                        term_months=int(tenure_months),
                        due_day=due_day,
                    )
                    db.add(lib)
                    liabilities_map[(str(lender).strip().lower(), str(loan_type).strip().lower())] = lib.id
            await db.flush()

        # 5. Liability Payments
        ws = wb["Debt"] if "Debt" in wb.sheetnames else None
        if ws:
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                if not row or not any(row): continue
                payment_date = cls.to_date(row[1])
                loan_type = row[3]
                lender = row[4]
                amount = row[5]
                if not amount or not lender or not payment_date: continue

                key = (str(lender).strip().lower(), str(loan_type).strip().lower())
                liability_id = liabilities_map.get(key)
                if liability_id:
                    pmt = LiabilityPayment(
                        id=uuid.uuid4(),
                        liability_id=liability_id,
                        payment_date=payment_date,
                        amount=Decimal(str(amount)),
                    )
                    db.add(pmt)
            await db.flush()

        # 6. Income
        ws = wb["Income"] if "Income" in wb.sheetnames else None
        if ws:
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                if not row or not any(row): continue
                date_val = cls.to_date(row[1])
                source = row[2]
                desc = row[3]
                amount = row[4]
                if not amount or not source or not date_val: continue
                if str(source).strip().lower() == "lent/expense back": continue

                source_id = get_lookup_id("income_source", source)
                if source_id:
                    inc = Income(
                        id=uuid.uuid4(),
                        user_id=user_id,
                        date=date_val,
                        income_source_id=source_id,
                        account_id=account_icici_id,
                        description=str(desc) if desc else None,
                        amount=Decimal(str(amount)),
                    )
                    db.add(inc)
            await db.flush()

        # 7. Expenses
        ws = wb["Expenses"] if "Expenses" in wb.sheetnames else None
        if ws:
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                if not row or not any(row): continue
                date_val = cls.to_date(row[1])
                item = row[2]
                amount = row[3]
                category = row[4]
                if not amount or not item or not category or not date_val: continue
                if str(category).strip().lower() == "lent": continue

                cat_id = get_lookup_id("expense_category", category, default_val="Other")
                if cat_id:
                    exp = Expense(
                        id=uuid.uuid4(),
                        user_id=user_id,
                        purchase_date=date_val,
                        item=str(item),
                        amount=Decimal(str(amount)),
                        category_id=cat_id,
                        account_id=account_icici_id,
                    )
                    db.add(exp)
            await db.flush()

        # 8. Investments
        ws = wb["Investment"] if "Investment" in wb.sheetnames else None
        if ws:
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                if not row or not any(row): continue
                date_val = cls.to_date(row[1])
                inv_type = row[2]
                amount = row[3]
                if not amount or not inv_type or not date_val: continue

                inst_str = "ICICI"
                if str(inv_type).lower().startswith("mutual"):
                    inst_str = "Vanguard"

                type_id = get_lookup_id("investment_type", inv_type) or get_lookup_id("investment_type", "Others")
                inst_id = get_lookup_id("institution", inst_str)

                if type_id and inst_id:
                    inv = Investment(
                        id=uuid.uuid4(),
                        user_id=user_id,
                        date=date_val,
                        investment_type_id=type_id,
                        institution_id=inst_id,
                        amount=Decimal(str(amount)),
                    )
                    db.add(inv)
            await db.flush()

        await db.commit()
        return True
